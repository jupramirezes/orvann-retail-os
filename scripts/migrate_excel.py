"""Migra datos desde Control_Operativo_Orvann.xlsx a SQLite."""
import sqlite3
import os
import re
from datetime import datetime, date

import openpyxl

EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', 'data', 'Control_Operativo_Orvann.xlsx')
DB_PATH = os.path.join(os.path.dirname(__file__), '..', 'data', 'orvann.db')

# Socios
SOCIOS = ['JP', 'KATHE', 'ANDRES']
SOCIO_RENAME = {'MILE': 'ANDRES'}

TALLAS_CONOCIDAS = ['S', 'M', 'L', 'XL', '2XL']

CATEGORIAS_CONOCIDAS = [
    'Camisa', 'Hoodie', 'Chompa', 'Buzo', 'Chaqueta',
    'Jogger', 'Sudadera', 'Pantaloneta',
]


def fix_socio(nombre):
    """Renombra MILE -> ANDRES."""
    if nombre:
        return SOCIO_RENAME.get(nombre.strip(), nombre.strip())
    return nombre


def parse_producto(nombre):
    """Extrae categoria, talla y color del nombre del producto."""
    if not nombre:
        return None, None, None

    nombre = nombre.strip()
    categoria = None
    talla = None
    color = None

    # Detectar categoría
    nombre_lower = nombre.lower()
    if nombre_lower.startswith('camisa'):
        categoria = 'Camisa'
    elif nombre_lower.startswith('hoodie'):
        categoria = 'Hoodie'
    elif nombre_lower.startswith('chompa'):
        categoria = 'Chompa'
    elif nombre_lower.startswith('buzo'):
        categoria = 'Buzo'
    elif nombre_lower.startswith('chaqueta'):
        categoria = 'Chaqueta'
    elif nombre_lower.startswith('jogger'):
        categoria = 'Jogger'
    elif nombre_lower.startswith('sudadera'):
        categoria = 'Sudadera'
    elif nombre_lower.startswith('pantaloneta'):
        categoria = 'Pantaloneta'

    # Buscar talla en el nombre
    parts = nombre.split()
    talla_idx = None
    for i, part in enumerate(parts):
        if part.upper() in TALLAS_CONOCIDAS:
            talla = part.upper()
            talla_idx = i
            break

    # El color es todo lo que viene después de la talla
    if talla_idx is not None and talla_idx + 1 < len(parts):
        color = ' '.join(parts[talla_idx + 1:]).strip()
    elif talla_idx is None:
        # Sin talla, intentar extraer color del final
        # Para productos como "Camisa Réplica 1.1" no hay color
        pass

    # Limpiar color vacío
    if color == '':
        color = None

    return categoria, talla, color


def to_date_str(val):
    """Convierte valor de celda a string de fecha YYYY-MM-DD."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, date):
        return val.strftime('%Y-%m-%d')
    return str(val)


def migrate_productos(ws, conn):
    """Migra hoja Inventario -> tabla productos."""
    c = conn.cursor()
    count = 0
    total_stock = 0
    seen_skus = set()

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=False):
        vals = {cell.column_letter: cell.value for cell in row}
        sku = vals.get('A')
        if sku is None:
            break

        sku = str(sku).strip()
        nombre = str(vals.get('B', '')).strip()
        costo = float(vals.get('C', 0) or 0)
        precio_venta = float(vals.get('D', 0) or 0)
        stock = int(float(vals.get('G', 0) or 0))
        notas = vals.get('H')

        categoria, talla, color = parse_producto(nombre)

        # Manejar SKUs duplicados (ej: SUD-NEG-L aparece 3 veces)
        if sku in seen_skus:
            # Generar SKU único basado en el nombre real
            # Extraer la talla real del nombre para diferenciar
            _, real_talla, _ = parse_producto(nombre)
            if real_talla:
                # Reemplazar la talla en el SKU
                base_sku = sku.rsplit('-', 1)[0] if '-' in sku else sku
                new_sku = f"{base_sku}-{real_talla}"
                if new_sku in seen_skus:
                    new_sku = f"{sku}-{real_talla}"
                sku = new_sku

        seen_skus.add(sku)

        c.execute("""
            INSERT OR REPLACE INTO productos
            (sku, nombre, categoria, talla, color, costo, precio_venta, stock, stock_minimo, notas)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, 3, ?)
        """, (sku, nombre, categoria, talla, color, costo, precio_venta, stock, notas))

        count += 1
        total_stock += stock

    conn.commit()
    print(f"  Productos: {count} SKUs, {total_stock} unidades totales")
    return count, total_stock


def migrate_ventas(ws, conn):
    """Migra hoja Ventas -> tabla ventas + creditos_clientes."""
    c = conn.cursor()
    count = 0
    creditos = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        vals = {cell.column_letter: cell.value for cell in row}
        fecha = vals.get('A')
        if fecha is None:
            break

        sku = str(vals.get('B', '')).strip()
        precio_venta = float(vals.get('E', 0) or 0)
        metodo_pago = str(vals.get('F', '')).strip()
        cliente = vals.get('G')
        notas = vals.get('H')

        if cliente:
            cliente = str(cliente).strip()

        # Renombrar método de pago para consistencia
        if 'cr' in metodo_pago.lower():
            metodo_pago = 'Crédito'
        elif 'transf' in metodo_pago.lower():
            metodo_pago = 'Transferencia'
        elif 'dat' in metodo_pago.lower():
            metodo_pago = 'Datáfono'
        elif 'efect' in metodo_pago.lower():
            metodo_pago = 'Efectivo'

        fecha_str = to_date_str(fecha)
        total = precio_venta  # cantidad=1, descuento=0

        c.execute("""
            INSERT INTO ventas (fecha, sku, cantidad, precio_unitario, descuento_pct, total, metodo_pago, cliente, notas)
            VALUES (?, ?, 1, ?, 0, ?, ?, ?, ?)
        """, (fecha_str, sku, precio_venta, total, metodo_pago, cliente, notas))

        venta_id = c.lastrowid
        count += 1

        # Si es crédito, crear registro en creditos_clientes
        if metodo_pago == 'Crédito' and cliente:
            c.execute("""
                INSERT INTO creditos_clientes (venta_id, cliente, monto, fecha_credito, pagado, notas)
                VALUES (?, ?, ?, ?, 0, ?)
            """, (venta_id, cliente, total, fecha_str, notas))
            creditos += 1

    conn.commit()
    print(f"  Ventas: {count} registros, {creditos} créditos creados")
    return count, creditos


def migrate_gastos(ws, conn):
    """
    Migra hoja Gastos -> tabla gastos.
    Maneja gastos triplicados: agrupa por (fecha, descripción) y almacena
    el monto real (no triplicado) con detalle de quién pagó.
    """
    c = conn.cursor()

    # Leer todos los gastos crudos
    raw = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        vals = {cell.column_letter: cell.value for cell in row}
        fecha = vals.get('A')
        if fecha is None:
            break
        raw.append({
            'fecha': to_date_str(fecha),
            'categoria': str(vals.get('B', '')).strip(),
            'monto': float(vals.get('C', 0) or 0),
            'descripcion': str(vals.get('D', '')).strip(),
            'metodo_pago': str(vals.get('E', '')).strip() if vals.get('E') else None,
            'responsable': fix_socio(str(vals.get('F', '')).strip()) if vals.get('F') else None,
            'notas': str(vals.get('G', '')).strip() if vals.get('G') else None,
        })

    print(f"  Gastos crudos del Excel: {len(raw)} filas")

    # Agrupar por (fecha, descripcion) para detectar triplicados
    from collections import defaultdict
    groups = defaultdict(list)
    for g in raw:
        key = (g['fecha'], g['descripcion'])
        groups[key].append(g)

    count = 0
    for (fecha, descripcion), items in groups.items():
        categoria = items[0]['categoria']
        metodo_pago = items[0]['metodo_pago']

        # Determinar si es inversión: gastos dic 2025 - ene 2026
        es_inversion = 0
        if fecha:
            y, m = int(fecha[:4]), int(fecha[5:7])
            if (y == 2025 and m == 12) or (y == 2026 and m == 1) or (y == 2025 and m in [1, 2]):
                es_inversion = 1

        if len(items) == 3:
            # Triplicado: verificar si los 3 montos son iguales
            montos = [it['monto'] for it in items]
            responsables = [it['responsable'] for it in items]

            if montos[0] == montos[1] == montos[2]:
                # Los 3 pagaron igual -> pagado_por = ORVANN
                monto = montos[0]
                pagado_por = 'ORVANN'
                notas_parts = []
                for it in items:
                    if it['notas'] and it['notas'] != 'None':
                        notas_parts.append(f"{it['responsable']}: {it['notas']}")
                notas = '; '.join(notas_parts) if notas_parts else None
            else:
                # Montos diferentes -> quien más puso
                monto = max(montos)
                max_idx = montos.index(monto)
                pagado_por = responsables[max_idx]
                monto = sum(montos)  # El total real es la suma de los 3
                notas_parts = []
                for it in items:
                    notas_parts.append(f"{it['responsable']}: ${it['monto']:,.0f}")
                    if it['notas'] and it['notas'] != 'None':
                        notas_parts[-1] += f" ({it['notas']})"
                notas = '; '.join(notas_parts)

            c.execute("""
                INSERT INTO gastos (fecha, categoria, monto, descripcion, metodo_pago, pagado_por, es_inversion, notas)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (fecha, categoria, monto, descripcion, metodo_pago, pagado_por, es_inversion, notas))
            count += 1
        else:
            # No triplicado: insertar cada gasto individualmente
            for it in items:
                pagado_por = it['responsable'] or 'ORVANN'
                notas = it['notas'] if it['notas'] != 'None' else None

                c.execute("""
                    INSERT INTO gastos (fecha, categoria, monto, descripcion, metodo_pago, pagado_por, es_inversion, notas)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (fecha, it['categoria'], it['monto'], it['descripcion'],
                      it['metodo_pago'], pagado_por, es_inversion, notas))
                count += 1

    conn.commit()
    print(f"  Gastos reales: {count} registros (de {len(raw)} filas crudas)")
    return count


def migrate_costos_fijos(ws, conn):
    """Migra hoja Costos Fijos -> tabla costos_fijos."""
    c = conn.cursor()
    count = 0
    total = 0

    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=False):
        vals = {cell.column_letter: cell.value for cell in row}
        concepto = vals.get('A')
        monto = vals.get('B')

        if concepto is None or monto is None:
            # Saltar filas vacías y "TOTAL"
            if concepto and 'total' in str(concepto).lower():
                continue
            if monto is None:
                continue

        concepto = str(concepto).strip()
        if concepto.upper() == 'TOTAL' or not concepto:
            continue

        monto = float(monto or 0)
        notas = str(vals.get('C', '')).strip() if vals.get('C') else None
        if notas == 'None':
            notas = None

        c.execute("""
            INSERT INTO costos_fijos (concepto, monto_mensual, activo, notas)
            VALUES (?, ?, 1, ?)
        """, (concepto, monto, notas))
        count += 1
        total += monto

    conn.commit()
    print(f"  Costos fijos: {count} rubros, total ${total:,.0f}")
    return count, total


def migrate_pedidos(ws, conn):
    """Migra hoja Pedidos Proveedores -> tabla pedidos_proveedores."""
    c = conn.cursor()
    count = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        vals = {cell.column_letter: cell.value for cell in row}
        fecha = vals.get('A')
        if fecha is None:
            break

        proveedor = str(vals.get('B', '')).strip()
        descripcion = str(vals.get('C', '')).strip()
        unidades = int(float(vals.get('D', 0) or 0))
        costo_unitario = float(vals.get('E', 0) or 0)
        total = float(vals.get('F', 0) or 0)
        estado = str(vals.get('G', 'Pendiente')).strip()
        fecha_pago = to_date_str(vals.get('H'))
        notas = str(vals.get('I', '')).strip() if vals.get('I') else None
        if notas == 'None':
            notas = None

        # Renombrar MILE en notas
        if notas:
            notas = notas.replace('MILE', 'ANDRES')

        c.execute("""
            INSERT INTO pedidos_proveedores
            (fecha_pedido, proveedor, descripcion, unidades, costo_unitario, total, estado, fecha_entrega_est, notas)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (to_date_str(fecha), proveedor, descripcion, unidades, costo_unitario, total, estado, fecha_pago, notas))
        count += 1

    conn.commit()
    print(f"  Pedidos: {count} registros")
    return count


def run_migration(excel_path=None, db_path=None):
    """Ejecuta la migración completa."""
    if excel_path is None:
        excel_path = EXCEL_PATH
    if db_path is None:
        db_path = DB_PATH

    print(f"Migrando desde: {os.path.abspath(excel_path)}")
    print(f"Base de datos: {os.path.abspath(db_path)}")
    print()

    # Crear tablas primero
    from scripts.create_db import create_tables
    create_tables(db_path)

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    conn = sqlite3.connect(db_path)

    try:
        print("Migrando datos...")
        print()

        # 1. Productos
        ws = wb['Inventario']
        n_prod, total_stock = migrate_productos(ws, conn)

        # 2. Ventas
        ws = wb['Ventas']
        n_ventas, n_creditos = migrate_ventas(ws, conn)

        # 3. Gastos
        ws = wb['Gastos']
        n_gastos = migrate_gastos(ws, conn)

        # 4. Costos fijos
        ws = wb['Costos Fijos']
        n_cf, total_cf = migrate_costos_fijos(ws, conn)

        # 5. Pedidos
        ws = wb['Pedidos Proveedores']
        n_pedidos = migrate_pedidos(ws, conn)

        # Resumen
        print()
        print("=" * 50)
        print("RESUMEN DE MIGRACIÓN")
        print("=" * 50)
        print(f"  Productos:     {n_prod} SKUs ({total_stock} unidades)")
        print(f"  Ventas:        {n_ventas} registros")
        print(f"  Créditos:      {n_creditos} registros")
        print(f"  Gastos:        {n_gastos} registros reales")
        print(f"  Costos fijos:  {n_cf} rubros (${total_cf:,.0f}/mes)")
        print(f"  Pedidos:       {n_pedidos} registros")
        print("=" * 50)

        return {
            'productos': n_prod,
            'stock_total': total_stock,
            'ventas': n_ventas,
            'creditos': n_creditos,
            'gastos': n_gastos,
            'costos_fijos': n_cf,
            'total_cf': total_cf,
            'pedidos': n_pedidos,
        }
    finally:
        conn.close()
        wb.close()


if __name__ == '__main__':
    import sys
    sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
    run_migration()
