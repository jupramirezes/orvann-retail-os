"""Sincroniza datos del Excel actualizado con la BD existente.

Solo agrega datos nuevos, NO duplica. Seguro de ejecutar multiples veces.
Diferencias encontradas (Excel vs BD):
- 1 gasto nuevo: 2026-02-15 Imprevistos $8,500 JP
- 1 costo fijo nuevo: "Persona punto de venta" $0
- Excel usa "MILE" que corresponde a "ANDRES" en la BD
"""
import openpyxl
import sqlite3
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', 'Control_Operativo_Orvann.xlsx')
# Fallback to data/ folder
if not os.path.exists(EXCEL_PATH):
    EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', 'data', 'Control_Operativo_Orvann.xlsx')

DB_PATH = os.path.join(os.path.dirname(__file__), '..', 'data', 'orvann.db')

# Mapeo de nombres del Excel a la BD
NOMBRE_MAP = {'MILE': 'ANDRES'}


def sync(db_path=None):
    if db_path is None:
        db_path = DB_PATH

    if not os.path.exists(EXCEL_PATH):
        print(f"ERROR: Excel no encontrado en {EXCEL_PATH}")
        return

    if not os.path.exists(db_path):
        print(f"ERROR: BD no encontrada en {db_path}")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row

    changes = 0

    # ── 1. Sync Gastos ─────────────────────────────────────
    print("=== GASTOS ===")
    ws = wb['Gastos']
    excel_gastos = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row and row[0] and str(row[0]) != 'None' and row[2]:
            fecha = str(row[0])[:10]
            categoria = str(row[1]) if row[1] else ''
            monto = float(row[2]) if row[2] else 0
            descripcion = str(row[3]) if row[3] and str(row[3]) != 'None' else ''
            metodo = str(row[4]) if row[4] and str(row[4]) != 'None' else ''
            responsable = str(row[5]) if row[5] else ''
            # Mapear nombres
            responsable = NOMBRE_MAP.get(responsable, responsable)
            if monto > 0:
                excel_gastos.append({
                    'fecha': fecha, 'categoria': categoria, 'monto': monto,
                    'descripcion': descripcion, 'metodo': metodo,
                    'responsable': responsable,
                })

    # Comparar con BD
    db_gastos = conn.execute("SELECT * FROM gastos ORDER BY fecha, id").fetchall()
    db_set = set()
    for g in db_gastos:
        key = (g['fecha'], g['pagado_por'], round(g['monto']))
        db_set.add(key)

    new_gastos = []
    for g in excel_gastos:
        key = (g['fecha'], g['responsable'], round(g['monto']))
        if key not in db_set:
            new_gastos.append(g)

    if new_gastos:
        for g in new_gastos:
            conn.execute("""
                INSERT INTO gastos (fecha, categoria, monto, descripcion, metodo_pago, pagado_por, es_inversion)
                VALUES (?, ?, ?, ?, ?, ?, 0)
            """, (g['fecha'], g['categoria'], g['monto'], g['descripcion'],
                  g['metodo'] or None, g['responsable']))
            print(f"  + Gasto: {g['fecha']} | {g['categoria']} | ${g['monto']:,.0f} | {g['descripcion']} | {g['responsable']}")
            changes += 1
    else:
        print("  (sin gastos nuevos)")

    # ── 2. Sync Costos Fijos ───────────────────────────────
    print("\n=== COSTOS FIJOS ===")
    ws = wb['Costos Fijos']
    excel_costos = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        if row and row[0] and str(row[0]).strip() and str(row[0]) != 'None':
            concepto = str(row[0]).strip()
            monto = float(row[1]) if row[1] else 0
            notas = str(row[2]) if row[2] and str(row[2]) != 'None' else ''
            if concepto not in ('TOTAL', 'TOTAL:'):
                excel_costos.append({'concepto': concepto, 'monto': monto, 'notas': notas})

    db_costos = conn.execute("SELECT * FROM costos_fijos").fetchall()
    db_conceptos = {c['concepto'].lower(): dict(c) for c in db_costos}

    for ec in excel_costos:
        found = False
        for key, dc in db_conceptos.items():
            if key.startswith(ec['concepto'].lower()[:10]) or ec['concepto'].lower().startswith(key[:10]):
                found = True
                # Check if monto changed
                if abs(dc['monto_mensual'] - ec['monto']) > 1:
                    conn.execute("UPDATE costos_fijos SET monto_mensual = ? WHERE id = ?",
                                 (ec['monto'], dc['id']))
                    print(f"  ~ Actualizado: {ec['concepto']} ${dc['monto_mensual']:,.0f} -> ${ec['monto']:,.0f}")
                    changes += 1
                break
        if not found:
            conn.execute("""
                INSERT INTO costos_fijos (concepto, monto_mensual, activo, notas)
                VALUES (?, ?, 1, ?)
            """, (ec['concepto'], ec['monto'], ec['notas'] or None))
            print(f"  + Nuevo: {ec['concepto']} = ${ec['monto']:,.0f}")
            changes += 1

    if changes == 0:
        print("  (sin cambios en costos fijos)")

    # ── 3. Sync precios de productos ───────────────────────
    print("\n=== PRODUCTOS (precios) ===")
    ws = wb['Inventario']
    price_changes = 0
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        if row and row[0] and str(row[0]).strip() and str(row[0]) != 'None':
            sku = str(row[0]).strip()
            costo = float(row[2]) if row[2] else 0
            precio = float(row[3]) if row[3] else 0

            db_prod = conn.execute("SELECT * FROM productos WHERE sku = ?", (sku,)).fetchone()
            if db_prod:
                if abs(db_prod['costo'] - costo) > 1 or abs(db_prod['precio_venta'] - precio) > 1:
                    conn.execute("UPDATE productos SET costo = ?, precio_venta = ? WHERE sku = ?",
                                 (costo, precio, sku))
                    print(f"  ~ {sku}: costo ${db_prod['costo']:,.0f}->${costo:,.0f} | precio ${db_prod['precio_venta']:,.0f}->${precio:,.0f}")
                    price_changes += 1
                    changes += 1

    if price_changes == 0:
        print("  (sin cambios de precios)")

    # ── Commit ─────────────────────────────────────────────
    conn.commit()
    conn.close()

    print(f"\n{'='*40}")
    print(f"TOTAL CAMBIOS: {changes}")
    print("Sync completado.")


if __name__ == '__main__':
    sync()
