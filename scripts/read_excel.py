"""Script temporal para leer el Excel actualizado y comparar con BD."""
import openpyxl
import sqlite3
import os
import sys

# Fix encoding for Windows console
sys.stdout.reconfigure(encoding='utf-8')

EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', 'Control_Operativo_Orvann.xlsx')
DB_PATH = os.path.join(os.path.dirname(__file__), '..', 'data', 'orvann.db')

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

# === INVENTARIO ===
print("=" * 60)
print("INVENTARIO (Excel)")
print("=" * 60)
ws = wb['Inventario']
# Header at row 4: SKU, Producto, Costo, Precio Venta, Margen %, Entradas, Stock Actual, Notas
excel_products = []
for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
    if row and row[0] and str(row[0]).strip() and str(row[0]) != 'None':
        sku = str(row[0]).strip()
        nombre = str(row[1]).strip() if row[1] else ''
        costo = float(row[2]) if row[2] else 0
        precio = float(row[3]) if row[3] else 0
        stock = int(row[6]) if row[6] else 0  # Stock Actual column
        notas = str(row[7]) if row[7] and str(row[7]) != 'None' else ''
        excel_products.append({
            'sku': sku, 'nombre': nombre, 'costo': costo,
            'precio': precio, 'stock': stock, 'notas': notas
        })

print("Total products in Excel:", len(excel_products))
print("Total stock in Excel:", sum(p['stock'] for p in excel_products))
# Show first 5
for p in excel_products[:5]:
    print("  {} | {} | cost={} | price={} | stock={}".format(
        p['sku'], p['nombre'][:40], p['costo'], p['precio'], p['stock']))
print("  ...")

# === VENTAS ===
print("\n" + "=" * 60)
print("VENTAS (Excel)")
print("=" * 60)
ws = wb['Ventas']
# Header at row 1
excel_ventas = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row and row[0] and str(row[0]) != 'None' and row[1] and str(row[1]) != 'None':
        fecha = str(row[0])[:10] if row[0] else ''
        sku = str(row[1]).strip() if row[1] else ''
        producto = str(row[2]).strip() if row[2] else ''
        costo = row[3]
        precio = row[4]
        metodo = str(row[5]) if row[5] else ''
        cliente = str(row[6]) if row[6] and str(row[6]) != 'None' else ''
        notas = str(row[7]) if row[7] and str(row[7]) != 'None' else ''
        # Filter out summary rows
        if sku and not sku.startswith('Total') and not sku.startswith('Ingres') and not sku.startswith('Costo') and not sku.startswith('Margen'):
            excel_ventas.append({
                'fecha': fecha, 'sku': sku, 'producto': producto,
                'costo': costo, 'precio': precio, 'metodo': metodo,
                'cliente': cliente, 'notas': notas
            })

print("Total ventas in Excel:", len(excel_ventas))
for v in excel_ventas:
    print("  {} | {} | {} | ${} | {} | {}".format(
        v['fecha'], v['sku'], v['producto'][:30], v['precio'], v['metodo'], v['cliente']))

# === GASTOS ===
print("\n" + "=" * 60)
print("GASTOS (Excel)")
print("=" * 60)
ws = wb['Gastos']
excel_gastos = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row and row[0] and str(row[0]) != 'None' and row[2]:
        fecha = str(row[0])[:10] if row[0] else ''
        categoria = str(row[1]) if row[1] else ''
        monto = float(row[2]) if row[2] else 0
        descripcion = str(row[3]) if row[3] and str(row[3]) != 'None' else ''
        metodo = str(row[4]) if row[4] and str(row[4]) != 'None' else ''
        responsable = str(row[5]) if row[5] else ''
        notas = str(row[6]) if row[6] and str(row[6]) != 'None' else ''
        if monto > 0:
            excel_gastos.append({
                'fecha': fecha, 'categoria': categoria, 'monto': monto,
                'descripcion': descripcion, 'metodo': metodo,
                'responsable': responsable, 'notas': notas
            })

print("Total gastos in Excel:", len(excel_gastos))
by_socio = {}
for g in excel_gastos:
    r = g['responsable']
    by_socio[r] = by_socio.get(r, 0) + g['monto']
for s, t in sorted(by_socio.items()):
    print("  {}: ${:,.0f}".format(s, t))
total_excel = sum(g['monto'] for g in excel_gastos)
print("  TOTAL: ${:,.0f}".format(total_excel))

# Last 15 gastos
print("\nAll gastos sorted by date:")
excel_gastos_sorted = sorted(excel_gastos, key=lambda x: x['fecha'])
for g in excel_gastos_sorted:
    print("  {} | {} | ${:,.0f} | {} | {}".format(
        g['fecha'], g['categoria'][:20], g['monto'], g['descripcion'][:35], g['responsable']))

# === PEDIDOS ===
print("\n" + "=" * 60)
print("PEDIDOS (Excel)")
print("=" * 60)
ws = wb['Pedidos Proveedores']
excel_pedidos = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row and row[0] and str(row[0]) != 'None' and row[1]:
        fecha = str(row[0])[:10] if row[0] else ''
        proveedor = str(row[1]).strip() if row[1] else ''
        desc = str(row[2]).strip() if row[2] else ''
        # Filter out summary rows
        if proveedor and not proveedor.startswith('TOTAL'):
            unidades = int(float(row[3])) if row[3] else 0
            costo_u = float(row[4]) if row[4] else 0
            total = float(row[5]) if row[5] else 0
            estado = str(row[6]) if row[6] else ''
            fecha_pago = str(row[7])[:10] if row[7] and str(row[7]) != 'None' else ''
            notas = str(row[8]) if row[8] and str(row[8]) != 'None' else ''
            excel_pedidos.append({
                'fecha': fecha, 'proveedor': proveedor, 'descripcion': desc,
                'unidades': unidades, 'costo_u': costo_u, 'total': total,
                'estado': estado, 'fecha_pago': fecha_pago, 'notas': notas
            })

print("Total pedidos in Excel:", len(excel_pedidos))
for p in excel_pedidos:
    print("  {} | {} | {} | {}x${:,.0f} = ${:,.0f} | {}".format(
        p['fecha'], p['proveedor'], p['descripcion'][:30],
        p['unidades'], p['costo_u'], p['total'], p['estado']))

# === COSTOS FIJOS ===
print("\n" + "=" * 60)
print("COSTOS FIJOS (Excel)")
print("=" * 60)
ws = wb['Costos Fijos']
excel_costos = []
for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
    if row and row[0] and str(row[0]).strip() and str(row[0]) != 'None':
        concepto = str(row[0]).strip()
        monto = float(row[1]) if row[1] else 0
        notas = str(row[2]) if row[2] and str(row[2]) != 'None' else ''
        if concepto != 'TOTAL' and concepto != 'TOTAL:':
            excel_costos.append({'concepto': concepto, 'monto': monto, 'notas': notas})

print("Total costos fijos in Excel:", len(excel_costos))
for c in excel_costos:
    print("  {}: ${:,.0f}{}".format(c['concepto'], c['monto'], ' | ' + c['notas'] if c['notas'] else ''))
print("  TOTAL: ${:,.0f}/mes".format(sum(c['monto'] for c in excel_costos)))

# === COMPARE WITH BD ===
print("\n" + "=" * 60)
print("COMPARACION CON BD")
print("=" * 60)
if os.path.exists(DB_PATH):
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    # Products
    db_prods = conn.execute("SELECT * FROM productos").fetchall()
    db_skus = {r['sku']: dict(r) for r in db_prods}
    excel_skus = {p['sku']: p for p in excel_products}

    print("\nProductos: BD={}, Excel={}".format(len(db_prods), len(excel_products)))
    new = set(excel_skus.keys()) - set(db_skus.keys())
    removed = set(db_skus.keys()) - set(excel_skus.keys())
    if new:
        print("  NUEVOS en Excel:")
        for s in new:
            p = excel_skus[s]
            print("    {} | {} | cost={} | price={}".format(s, p['nombre'][:40], p['costo'], p['precio']))
    if removed:
        print("  En BD pero NO en Excel:")
        for s in removed:
            p = db_skus[s]
            print("    {} | {} | stock={}".format(s, p['nombre'][:40], p['stock']))

    # Price changes
    print("\n  Diferencias de costo/precio:")
    changes = 0
    for sku in set(excel_skus.keys()) & set(db_skus.keys()):
        ep = excel_skus[sku]
        dp = db_skus[sku]
        if ep['costo'] != dp['costo'] or ep['precio'] != dp['precio_venta']:
            changes += 1
            print("    {} costo: BD={} Excel={} | precio: BD={} Excel={}".format(
                sku, dp['costo'], ep['costo'], dp['precio_venta'], ep['precio']))
    if changes == 0:
        print("    (ninguna diferencia)")

    # Gastos
    db_gastos = conn.execute("SELECT * FROM gastos ORDER BY fecha, id").fetchall()
    print("\nGastos: BD={}, Excel={}".format(len(db_gastos), len(excel_gastos)))
    by_socio_db = {}
    for g in db_gastos:
        s = g['pagado_por']
        by_socio_db[s] = by_socio_db.get(s, 0) + g['monto']
    for s, t in sorted(by_socio_db.items()):
        excel_t = by_socio.get(s, 0)
        diff = excel_t - t
        status = " (OK)" if abs(diff) < 1 else " DIFF={:+,.0f}".format(diff)
        print("  {}: BD=${:,.0f} | Excel=${:,.0f}{}".format(s, t, excel_t, status))

    # Ventas
    db_ventas = conn.execute("SELECT * FROM ventas").fetchall()
    print("\nVentas: BD={}, Excel={}".format(len(db_ventas), len(excel_ventas)))

    # Pedidos
    db_pedidos = conn.execute("SELECT * FROM pedidos_proveedores ORDER BY fecha_pedido").fetchall()
    print("\nPedidos: BD={}, Excel={}".format(len(db_pedidos), len(excel_pedidos)))

    # Costos fijos
    db_costos = conn.execute("SELECT * FROM costos_fijos").fetchall()
    print("\nCostos fijos: BD={}, Excel={}".format(len(db_costos), len(excel_costos)))
    for ec in excel_costos:
        found = False
        for dc in db_costos:
            if dc['concepto'] == ec['concepto'] or dc['concepto'].lower().startswith(ec['concepto'].lower()[:10]):
                found = True
                if abs(dc['monto_mensual'] - ec['monto']) > 1:
                    print("  {}: BD=${:,.0f} -> Excel=${:,.0f}".format(ec['concepto'], dc['monto_mensual'], ec['monto']))
                break
        if not found:
            print("  NUEVO: {} = ${:,.0f}".format(ec['concepto'], ec['monto']))

    conn.close()
else:
    print("BD no encontrada en", DB_PATH)
